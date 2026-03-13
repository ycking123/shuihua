from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import or_
from sqlalchemy.orm import Session

from ..models import Meeting, MeetingRoom, SysUser, Todo


ROOM_SOURCE_PRIORITY = [
    "会议室列表(4).xlsx",
    "帝王会议室列表.xlsx",
    "会议室列表.xlsx",
]


class MeetingServiceError(Exception):
    """会议域统一异常。"""

    def __init__(
        self,
        message: str,
        status_code: int = 400,
        code: Optional[str] = None,
        extra: Optional[Dict[str, Any]] = None,
    ) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.code = code
        self.extra = extra or {}


class MeetingService:
    """会议域统一业务服务。"""

    def __init__(self) -> None:
        self.project_root = Path(__file__).resolve().parents[2]
        self.xlsx_dir = self.project_root / "xlsx"

    def list_meetings(
        self,
        db: Session,
        skip: int = 0,
        limit: int = 100,
        sort_by: str = "start_time",
        status: Optional[str] = None,
        keyword: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        query = db.query(Meeting)
        if status:
            query = query.filter(Meeting.status == status)
        if keyword:
            query = query.filter(Meeting.title.like(f"%{keyword}%"))

        if sort_by in {"created_at", "meeting_created_at"}:
            query = query.order_by(Meeting.created_at.desc())
        else:
            query = query.order_by(Meeting.start_time.desc())

        meetings = query.offset(skip).limit(limit).all()
        return [self.serialize_meeting(db, meeting) for meeting in meetings]

    def get_meeting(self, db: Session, meeting_id: str) -> Dict[str, Any]:
        meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
        if not meeting:
            raise MeetingServiceError("会议不存在", status_code=404)
        return self.serialize_meeting(db, meeting)

    def import_link(
        self,
        db: Session,
        payload: Dict[str, Any],
        current_user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        meeting_url = self.normalize_text(payload.get("meeting_url") or payload.get("url"))
        if not meeting_url:
            raise MeetingServiceError("缺少会议链接", code="missing_meeting_url")

        title = self.normalize_text(payload.get("title")) or "链接导入会议"
        start_time = self.parse_datetime_value(payload.get("start_time")) or datetime.now()
        end_time = self.parse_datetime_value(payload.get("end_time")) or (start_time + timedelta(minutes=60))

        if payload.get("meeting_id"):
            return self.update_meeting(
                db,
                meeting_id=str(payload["meeting_id"]),
                payload={
                    "title": title,
                    "meeting_url": meeting_url,
                    "description": payload.get("description"),
                    "summary": payload.get("summary") or "已导入会议链接，待后续解析纪要。",
                    "start_time": start_time,
                    "end_time": end_time,
                    "source_type": "link_import",
                    "location": payload.get("location"),
                },
                current_user_id=current_user_id,
            )

        return self.create_meeting(
            db,
            payload={
                "title": title,
                "start_time": start_time,
                "end_time": end_time,
                "location": self.normalize_text(payload.get("location")) or "待补充",
                "meeting_url": meeting_url,
                "description": payload.get("description"),
                "summary": payload.get("summary") or "已导入会议链接，待后续解析纪要。",
                "source_type": "link_import",
                "channel": payload.get("channel") or "web",
                "participants": payload.get("participants"),
            },
            current_user_id=current_user_id,
            default_source_type="link_import",
        )

    def transcribe_meeting(
        self,
        db: Session,
        payload: Dict[str, Any],
        current_user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        transcript = self.normalize_text(payload.get("transcript") or payload.get("text"))
        if not transcript:
            raise MeetingServiceError("缺少转写文本", code="missing_transcript")

        summary = self.normalize_text(payload.get("summary")) or self.build_summary(transcript)
        if payload.get("meeting_id"):
            return self.update_meeting(
                db,
                meeting_id=str(payload["meeting_id"]),
                payload={
                    "transcript": transcript,
                    "summary": summary,
                    "audio_file_key": payload.get("audio_file_key"),
                    "source_type": "recording",
                    "sync_status": payload.get("sync_status") or "none",
                },
                current_user_id=current_user_id,
            )

        title = self.normalize_text(payload.get("title")) or self.derive_title_from_text(transcript, default="录音纪要")
        start_time = self.parse_datetime_value(payload.get("start_time")) or datetime.now()
        end_time = self.parse_datetime_value(payload.get("end_time")) or (start_time + timedelta(minutes=60))
        return self.create_meeting(
            db,
            payload={
                "title": title,
                "start_time": start_time,
                "end_time": end_time,
                "location": self.normalize_text(payload.get("location")) or "待补充",
                "participants": payload.get("participants"),
                "transcript": transcript,
                "summary": summary,
                "audio_file_key": payload.get("audio_file_key"),
                "description": payload.get("description"),
                "source_type": "recording",
                "channel": payload.get("channel") or "web",
                "sync_status": payload.get("sync_status") or "none",
            },
            current_user_id=current_user_id,
            default_source_type="recording",
        )

    def generate_summary(self, db: Session, meeting_id: str) -> Dict[str, Any]:
        meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
        if not meeting:
            raise MeetingServiceError("会议不存在", status_code=404)

        source_text = self.normalize_text(meeting.transcript) or self.normalize_text(meeting.description)
        if not source_text:
            raise MeetingServiceError("当前会议缺少可摘要内容")

        meeting.summary = self.build_summary(source_text)
        db.commit()
        db.refresh(meeting)
        return self.serialize_meeting(db, meeting)

    def create_meeting(
        self,
        db: Session,
        payload: Dict[str, Any],
        current_user_id: Optional[str] = None,
        default_channel: str = "web",
        default_source_type: str = "manual",
    ) -> Dict[str, Any]:
        self.ensure_room_seeded_from_xlsx(db)

        title = self.normalize_text(payload.get("title"))
        if not title:
            raise MeetingServiceError("缺少会议主题", code="missing_title")

        start_time = self.parse_datetime_value(payload.get("start_time"))
        if not start_time:
            raise MeetingServiceError("缺少会议时间", code="missing_start_time")

        end_time = self.parse_datetime_value(payload.get("end_time"))
        if not end_time:
            duration_minutes = self.parse_duration_minutes(payload.get("duration"))
            end_time = start_time + timedelta(minutes=duration_minutes or 60)

        if end_time <= start_time:
            raise MeetingServiceError("结束时间必须晚于开始时间")

        room = self.resolve_room(
            db=db,
            room_id=payload.get("room_id"),
            room_name=payload.get("room_name"),
            site_name=payload.get("site_name"),
            allow_external=bool(payload.get("location")) and not bool(payload.get("room_name")),
        )
        raw_location = self.normalize_text(payload.get("location"))
        location = raw_location
        if room and (not raw_location or raw_location == room.room_name):
            location = room.location_text
        if not room and not location:
            raise MeetingServiceError("缺少会议室或地点信息", code="missing_location")

        if room:
            self.validate_room_conflict(db, room.id, start_time, end_time)

        transcript = self.normalize_text(payload.get("transcript"))
        description = self.normalize_text(payload.get("description"))
        summary = self.normalize_text(payload.get("summary")) or self.build_summary(transcript or description or "")
        participants = self.normalize_people(payload.get("participants"))
        organizer_user_id = self.parse_user_id(current_user_id)

        meeting = Meeting(
            title=title,
            description=description,
            organizer_user_id=organizer_user_id,
            organizer_id=str(current_user_id) if current_user_id else None,
            status=self.normalize_text(payload.get("status")) or "scheduled",
            channel=self.normalize_text(payload.get("channel")) or default_channel,
            source_type=self.normalize_text(payload.get("source_type")) or default_source_type,
            start_time=start_time,
            end_time=end_time,
            room_id=room.id if room else None,
            location=location,
            participants_json=participants,
            meeting_url=self.normalize_text(payload.get("meeting_url")),
            wecom_schedule_id=self.normalize_text(payload.get("wecom_schedule_id")),
            audio_file_key=self.normalize_text(payload.get("audio_file_key")),
            summary=summary,
            transcript=transcript,
            sync_status=self.normalize_text(payload.get("sync_status")) or "none",
        )

        db.add(meeting)
        db.commit()
        db.refresh(meeting)
        return self.serialize_meeting(db, meeting)

    def update_meeting(
        self,
        db: Session,
        meeting_id: str,
        payload: Dict[str, Any],
        current_user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
        if not meeting:
            raise MeetingServiceError("会议不存在", status_code=404)

        self.ensure_room_seeded_from_xlsx(db)

        if "title" in payload:
            title = self.normalize_text(payload.get("title"))
            if title:
                meeting.title = title

        if "description" in payload:
            meeting.description = self.normalize_text(payload.get("description"))

        if "status" in payload:
            meeting.status = self.normalize_text(payload.get("status")) or meeting.status

        if "channel" in payload:
            meeting.channel = self.normalize_text(payload.get("channel")) or meeting.channel

        if "source_type" in payload:
            meeting.source_type = self.normalize_text(payload.get("source_type")) or meeting.source_type

        if "meeting_url" in payload:
            meeting.meeting_url = self.normalize_text(payload.get("meeting_url"))

        if "audio_file_key" in payload:
            meeting.audio_file_key = self.normalize_text(payload.get("audio_file_key"))

        if "sync_status" in payload:
            meeting.sync_status = self.normalize_text(payload.get("sync_status")) or meeting.sync_status

        if "wecom_schedule_id" in payload:
            meeting.wecom_schedule_id = self.normalize_text(payload.get("wecom_schedule_id"))

        if "participants" in payload:
            meeting.participants_json = self.normalize_people(payload.get("participants"))

        if "transcript" in payload:
            meeting.transcript = self.normalize_text(payload.get("transcript"))

        if "summary" in payload:
            meeting.summary = self.normalize_text(payload.get("summary"))

        start_time = self.parse_datetime_value(payload.get("start_time")) if "start_time" in payload else meeting.start_time
        end_time = self.parse_datetime_value(payload.get("end_time")) if "end_time" in payload else meeting.end_time
        if not end_time and "duration" in payload:
            duration_minutes = self.parse_duration_minutes(payload.get("duration"))
            end_time = start_time + timedelta(minutes=duration_minutes or 60)

        room = None
        room_fields_changed = any(key in payload for key in ["room_id", "room_name", "site_name", "location"])
        if room_fields_changed:
            current_room = db.query(MeetingRoom).filter(MeetingRoom.id == meeting.room_id).first() if meeting.room_id else None
            explicit_location = self.normalize_text(payload.get("location")) if "location" in payload else None
            use_external_location = (
                "location" in payload
                and explicit_location is not None
                and "room_id" not in payload
                and "room_name" not in payload
            )
            room = self.resolve_room(
                db=db,
                room_id=None if use_external_location else (
                    payload.get("room_id") if "room_id" in payload else (current_room.id if current_room else None)
                ),
                room_name=None if use_external_location else (
                    payload.get("room_name") if "room_name" in payload else (current_room.room_name if current_room else None)
                ),
                site_name=payload.get("site_name") if "site_name" in payload else (current_room.site_name if current_room else None),
                allow_external=use_external_location,
            )
        elif meeting.room_id:
            room = db.query(MeetingRoom).filter(MeetingRoom.id == meeting.room_id).first()

        if room:
            self.validate_room_conflict(db, room.id, start_time, end_time, exclude_meeting_id=meeting.id)
            meeting.room_id = room.id
            raw_location = self.normalize_text(payload.get("location"))
            meeting.location = room.location_text if not raw_location or raw_location == room.room_name else raw_location
        elif "location" in payload:
            meeting.room_id = None
            meeting.location = self.normalize_text(payload.get("location"))

        if start_time and end_time and end_time <= start_time:
            raise MeetingServiceError("结束时间必须晚于开始时间")

        meeting.start_time = start_time
        meeting.end_time = end_time

        if current_user_id and not meeting.organizer_user_id:
            meeting.organizer_user_id = self.parse_user_id(current_user_id)
            meeting.organizer_id = str(current_user_id)

        if not meeting.summary and meeting.transcript:
            meeting.summary = self.build_summary(meeting.transcript)

        db.commit()
        db.refresh(meeting)
        return self.serialize_meeting(db, meeting)

    def cancel_meeting(self, db: Session, meeting_id: str) -> Dict[str, Any]:
        meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
        if not meeting:
            raise MeetingServiceError("会议不存在", status_code=404)

        meeting.status = "cancelled"
        db.commit()
        db.refresh(meeting)
        return self.serialize_meeting(db, meeting)

    def get_meeting_todos(self, db: Session, meeting_id: str) -> List[Dict[str, Any]]:
        meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
        if not meeting:
            raise MeetingServiceError("会议不存在", status_code=404)

        todos = (
            db.query(Todo)
            .filter(Todo.source_message_id == meeting_id, Todo.is_deleted == False)
            .order_by(Todo.created_at.desc())
            .all()
        )
        return [
            {
                "id": todo.id,
                "title": todo.title,
                "content": todo.content,
                "priority": todo.priority,
                "status": todo.status,
                "assignee": todo.sender,
                "due_at": todo.due_at,
                "created_at": todo.created_at,
            }
            for todo in todos
        ]

    def list_rooms(
        self,
        db: Session,
        keyword: Optional[str] = None,
        site_name: Optional[str] = None,
        building_name: Optional[str] = None,
        floor_label: Optional[str] = None,
        is_enabled: Optional[bool] = None,
        min_capacity: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        self.ensure_room_seeded_from_xlsx(db)

        query = db.query(MeetingRoom)
        if keyword:
            query = query.filter(
                or_(
                    MeetingRoom.room_name.like(f"%{keyword}%"),
                    MeetingRoom.location_text.like(f"%{keyword}%"),
                    MeetingRoom.site_name.like(f"%{keyword}%"),
                )
            )
        if site_name:
            query = query.filter(MeetingRoom.site_name == site_name)
        if building_name:
            query = query.filter(MeetingRoom.building_name.like(f"%{building_name}%"))
        if floor_label:
            query = query.filter(MeetingRoom.floor_label.like(f"%{floor_label}%"))
        if is_enabled is not None:
            query = query.filter(MeetingRoom.is_enabled == is_enabled)
        if min_capacity is not None:
            query = query.filter(MeetingRoom.capacity >= min_capacity)

        rooms = (
            query.order_by(MeetingRoom.site_name.asc(), MeetingRoom.sort_order.asc(), MeetingRoom.room_name.asc()).all()
        )
        return [self.serialize_room(room) for room in rooms]

    def create_room(self, db: Session, payload: Dict[str, Any]) -> Dict[str, Any]:
        room_name = self.normalize_text(payload.get("room_name"))
        site_name = self.normalize_text(payload.get("site_name"))
        location_text = self.normalize_text(payload.get("location_text"))
        if not room_name or not site_name or not location_text:
            raise MeetingServiceError("会议室名称、园区归属和地点楼层不能为空")

        existing = (
            db.query(MeetingRoom)
            .filter(MeetingRoom.room_name == room_name, MeetingRoom.site_name == site_name)
            .first()
        )
        if existing:
            raise MeetingServiceError("同园区下已存在同名会议室")

        building_name, floor_label = self.extract_location_parts(location_text)
        manager_name = self.normalize_text(payload.get("manager_name"))
        room = MeetingRoom(
            room_name=room_name,
            site_name=site_name,
            location_text=location_text,
            building_name=self.normalize_text(payload.get("building_name")) or building_name,
            floor_label=self.normalize_text(payload.get("floor_label")) or floor_label,
            capacity=self.to_int(payload.get("capacity"), default=0),
            seat_layout=self.normalize_text(payload.get("seat_layout")),
            manager_name=manager_name,
            manager_user_id=self.resolve_manager_user_id(db, manager_name),
            sort_order=self.to_optional_int(payload.get("sort_order")),
            is_enabled=self.parse_bool(payload.get("is_enabled"), default=True),
            source_file=self.normalize_text(payload.get("source_file")),
            source_sheet=self.normalize_text(payload.get("source_sheet")),
        )
        db.add(room)
        db.commit()
        db.refresh(room)
        return self.serialize_room(room)

    def update_room(self, db: Session, room_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        room = db.query(MeetingRoom).filter(MeetingRoom.id == room_id).first()
        if not room:
            raise MeetingServiceError("会议室不存在", status_code=404)

        next_room_name = self.normalize_text(payload.get("room_name")) if "room_name" in payload else room.room_name
        next_site_name = self.normalize_text(payload.get("site_name")) if "site_name" in payload else room.site_name
        duplicate = (
            db.query(MeetingRoom)
            .filter(
                MeetingRoom.id != room_id,
                MeetingRoom.room_name == next_room_name,
                MeetingRoom.site_name == next_site_name,
            )
            .first()
        )
        if duplicate:
            raise MeetingServiceError("鍚屽洯鍖轰笅宸插瓨鍦ㄥ悓鍚嶄細璁")

        if "room_name" in payload:
            room.room_name = next_room_name or room.room_name
        if "site_name" in payload:
            room.site_name = next_site_name or room.site_name
        if "location_text" in payload:
            room.location_text = self.normalize_text(payload.get("location_text")) or room.location_text
            building_name, floor_label = self.extract_location_parts(room.location_text)
            room.building_name = self.normalize_text(payload.get("building_name")) or building_name
            room.floor_label = self.normalize_text(payload.get("floor_label")) or floor_label
        if "building_name" in payload and "location_text" not in payload:
            room.building_name = self.normalize_text(payload.get("building_name"))
        if "floor_label" in payload and "location_text" not in payload:
            room.floor_label = self.normalize_text(payload.get("floor_label"))
        if "capacity" in payload:
            room.capacity = self.to_int(payload.get("capacity"), default=room.capacity or 0)
        if "seat_layout" in payload:
            room.seat_layout = self.normalize_text(payload.get("seat_layout"))
        if "manager_name" in payload:
            room.manager_name = self.normalize_text(payload.get("manager_name"))
            room.manager_user_id = self.resolve_manager_user_id(db, room.manager_name)
        if "sort_order" in payload:
            room.sort_order = self.to_optional_int(payload.get("sort_order"))
        if "is_enabled" in payload:
            room.is_enabled = self.parse_bool(payload.get("is_enabled"), default=room.is_enabled)
        if "source_file" in payload:
            room.source_file = self.normalize_text(payload.get("source_file"))
        if "source_sheet" in payload:
            room.source_sheet = self.normalize_text(payload.get("source_sheet"))

        db.commit()
        db.refresh(room)
        return self.serialize_room(room)

    def delete_room(self, db: Session, room_id: str) -> Dict[str, Any]:
        room = db.query(MeetingRoom).filter(MeetingRoom.id == room_id).first()
        if not room:
            raise MeetingServiceError("会议室不存在", status_code=404)

        related_count = db.query(Meeting).filter(Meeting.room_id == room_id).count()
        if related_count > 0:
            room.is_enabled = False
            db.commit()
            return {"id": room_id, "deleted": False, "disabled": True, "message": "会议室已有关联会议，已自动停用"}

        db.delete(room)
        db.commit()
        return {"id": room_id, "deleted": True, "disabled": False}

    def get_room_usage(
        self,
        db: Session,
        usage_date: Optional[str] = None,
        room_id: Optional[str] = None,
        site_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        self.ensure_room_seeded_from_xlsx(db)

        target_date = self.parse_date_value(usage_date) or datetime.now().date()
        day_start = datetime.combine(target_date, time.min)
        day_end = day_start + timedelta(days=1)

        room_query = db.query(MeetingRoom)
        if room_id:
            room_query = room_query.filter(MeetingRoom.id == room_id)
        if site_name:
            room_query = room_query.filter(MeetingRoom.site_name == site_name)

        rooms = room_query.order_by(MeetingRoom.site_name.asc(), MeetingRoom.room_name.asc()).all()
        room_ids = [room.id for room in rooms]

        meetings = []
        if room_ids:
            meetings = (
                db.query(Meeting)
                .filter(
                    Meeting.room_id.in_(room_ids),
                    Meeting.status != "cancelled",
                    Meeting.start_time < day_end,
                    Meeting.end_time > day_start,
                )
                .order_by(Meeting.start_time.asc())
                .all()
            )

        usage_map: Dict[str, List[Meeting]] = {}
        for meeting in meetings:
            usage_map.setdefault(meeting.room_id, []).append(meeting)

        items = []
        for room in rooms:
            bookings = [
                {
                    "meeting_id": meeting.id,
                    "title": meeting.title,
                    "start_time": meeting.start_time,
                    "end_time": meeting.end_time,
                    "status": meeting.status,
                    "participants": self.normalize_people(meeting.participants_json),
                    "location": meeting.location,
                }
                for meeting in usage_map.get(room.id, [])
            ]
            items.append(
                {
                    "room": self.serialize_room(room),
                    "bookings": bookings,
                    "is_available": len(bookings) == 0,
                }
            )
        return {"date": target_date.isoformat(), "rooms": items}

    def execute_intent(
        self,
        db: Session,
        intent: str,
        params: Dict[str, Any],
        current_user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        if intent in {"meeting_create", "room_book"}:
            meeting = self.create_meeting(
                db,
                payload={
                    "title": params.get("title"),
                    "start_time": params.get("start_time"),
                    "end_time": params.get("end_time"),
                    "duration": params.get("duration"),
                    "participants": params.get("participants"),
                    "room_id": params.get("room_id"),
                    "room_name": params.get("room_name") or params.get("location"),
                    "site_name": params.get("site_name"),
                    "location": params.get("location"),
                    "description": params.get("description"),
                    "meeting_url": params.get("meeting_url"),
                    "source_type": "chat_intent",
                    "channel": "web",
                },
                current_user_id=current_user_id,
                default_source_type="chat_intent",
            )
            room_text = meeting.get("room_name") or meeting.get("location") or "待补充"
            return {
                "success": True,
                "message": f"已创建会议《{meeting['title']}》，时间 {meeting['start_time']}，地点 {room_text}。",
                "data": meeting,
            }

        if intent == "meeting_update":
            target = self.find_meeting_for_intent(db, params, current_user_id)
            updated = self.update_meeting(db, target.id, params, current_user_id=current_user_id)
            return {"success": True, "message": f"已更新会议《{updated['title']}》。", "data": updated}

        if intent == "meeting_cancel":
            target = self.find_meeting_for_intent(db, params, current_user_id)
            cancelled = self.cancel_meeting(db, target.id)
            return {"success": True, "message": f"已取消会议《{cancelled['title']}》。", "data": cancelled}

        if intent == "room_query":
            rooms = self.list_rooms(
                db,
                keyword=self.normalize_text(params.get("room_name") or params.get("keyword")),
                site_name=self.normalize_text(params.get("site_name")),
                min_capacity=self.to_optional_int(params.get("capacity")),
                is_enabled=True,
            )
            if not rooms:
                return {"success": True, "message": "没有找到匹配的会议室。", "data": {"rooms": []}}
            preview = [f"{room['site_name']}·{room['room_name']}（{room['capacity']}人）" for room in rooms[:5]]
            return {"success": True, "message": "可用会议室如下：\n" + "\n".join(preview), "data": {"rooms": rooms}}

        if intent == "meeting_query":
            meetings = self.list_meetings(
                db,
                limit=5,
                keyword=self.normalize_text(params.get("title")),
                status=self.normalize_text(params.get("status")),
            )
            if not meetings:
                return {"success": True, "message": "当前没有匹配的会议记录。", "data": {"meetings": []}}
            preview = [f"{item['title']}（{item['start_time']}）" for item in meetings[:5]]
            return {"success": True, "message": "找到以下会议：\n" + "\n".join(preview), "data": {"meetings": meetings}}

        if intent == "minutes_import":
            result = self.import_link(db, params, current_user_id=current_user_id)
            return {"success": True, "message": f"已导入会议链接《{result['title']}》。", "data": result}

        if intent == "minutes_record":
            result = self.transcribe_meeting(db, params, current_user_id=current_user_id)
            return {"success": True, "message": f"已保存录音纪要《{result['title']}》。", "data": result}

        raise MeetingServiceError("暂不支持的会议意图")

    def serialize_meeting(self, db: Session, meeting: Meeting) -> Dict[str, Any]:
        room = db.query(MeetingRoom).filter(MeetingRoom.id == meeting.room_id).first() if meeting.room_id else None
        todos_count = (
            db.query(Todo)
            .filter(Todo.source_message_id == meeting.id, Todo.is_deleted == False)
            .count()
        )
        organizer_value = (
            str(meeting.organizer_user_id)
            if meeting.organizer_user_id is not None
            else meeting.organizer_id
        )
        return {
            "id": meeting.id,
            "title": meeting.title,
            "description": meeting.description,
            "organizer_user_id": meeting.organizer_user_id,
            "organizer_id": organizer_value,
            "status": meeting.status,
            "channel": meeting.channel,
            "source_type": meeting.source_type,
            "start_time": meeting.start_time,
            "end_time": meeting.end_time,
            "created_at": meeting.created_at,
            "updated_at": meeting.updated_at,
            "location": meeting.location,
            "room_id": meeting.room_id,
            "room_name": room.room_name if room else None,
            "room_site_name": room.site_name if room else None,
            "meeting_url": meeting.meeting_url,
            "wecom_schedule_id": meeting.wecom_schedule_id,
            "audio_file_key": meeting.audio_file_key,
            "summary": meeting.summary,
            "transcript": meeting.transcript,
            "participants": self.normalize_people(meeting.participants_json),
            "participants_json": self.normalize_people(meeting.participants_json),
            "sync_status": meeting.sync_status,
            "todos_count": todos_count,
        }

    def serialize_room(self, room: MeetingRoom) -> Dict[str, Any]:
        return {
            "id": room.id,
            "room_name": room.room_name,
            "site_name": room.site_name,
            "location_text": room.location_text,
            "building_name": room.building_name,
            "floor_label": room.floor_label,
            "capacity": room.capacity,
            "seat_layout": room.seat_layout,
            "manager_name": room.manager_name,
            "manager_user_id": room.manager_user_id,
            "sort_order": room.sort_order,
            "is_enabled": room.is_enabled,
            "source_file": room.source_file,
            "source_sheet": room.source_sheet,
            "created_at": room.created_at,
            "updated_at": room.updated_at,
        }

    def validate_room_conflict(
        self,
        db: Session,
        room_id: str,
        start_time: datetime,
        end_time: datetime,
        exclude_meeting_id: Optional[str] = None,
    ) -> None:
        query = db.query(Meeting).filter(
            Meeting.room_id == room_id,
            Meeting.status != "cancelled",
            Meeting.start_time < end_time,
            Meeting.end_time > start_time,
        )
        if exclude_meeting_id:
            query = query.filter(Meeting.id != exclude_meeting_id)

        conflict = query.order_by(Meeting.start_time.asc()).first()
        if conflict:
            raise MeetingServiceError(
                f"会议室在 {conflict.start_time.strftime('%Y-%m-%d %H:%M')} 至 "
                f"{conflict.end_time.strftime('%H:%M')} 已被《{conflict.title}》占用",
                code="room_conflict",
            )

    def resolve_room(
        self,
        db: Session,
        room_id: Optional[str] = None,
        room_name: Optional[str] = None,
        site_name: Optional[str] = None,
        allow_external: bool = False,
    ) -> Optional[MeetingRoom]:
        if room_id:
            room = db.query(MeetingRoom).filter(MeetingRoom.id == room_id).first()
            if room:
                return room
            raise MeetingServiceError("指定的会议室不存在", code="room_not_found")

        room_name = self.normalize_text(room_name)
        if not room_name:
            return None

        query = db.query(MeetingRoom).filter(MeetingRoom.room_name == room_name)
        if site_name:
            query = query.filter(MeetingRoom.site_name == site_name)
        matches = query.order_by(MeetingRoom.site_name.asc()).all()
        if len(matches) == 1:
            return matches[0]

        if not matches:
            fuzzy_query = db.query(MeetingRoom).filter(MeetingRoom.room_name.like(f"%{room_name}%"))
            if site_name:
                fuzzy_query = fuzzy_query.filter(MeetingRoom.site_name == site_name)
            matches = fuzzy_query.order_by(MeetingRoom.site_name.asc()).all()

        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise MeetingServiceError(
                "会议室名称存在多个候选，请补充园区或更精确名称",
                code="room_ambiguous",
                extra={
                    "candidates": [
                        {
                            "id": item.id,
                            "room_name": item.room_name,
                            "site_name": item.site_name,
                            "location_text": item.location_text,
                        }
                        for item in matches[:10]
                    ]
                },
            )
        if allow_external:
            return None
        raise MeetingServiceError("未找到匹配的会议室", code="room_not_found")

    def ensure_room_seeded_from_xlsx(self, db: Session) -> int:
        room_count = db.query(MeetingRoom).count()
        if room_count > 0:
            return 0
        return self.import_rooms_from_xlsx(db)

    def import_rooms_from_xlsx(self, db: Session) -> int:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise MeetingServiceError(
                "缺少 openpyxl 依赖，无法导入会议室 Excel",
                status_code=500,
                code="missing_openpyxl",
            ) from exc

        records: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for file_name in ROOM_SOURCE_PRIORITY:
            file_path = self.xlsx_dir / file_name
            if not file_path.exists():
                continue

            workbook = load_workbook(file_path, data_only=True)
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue

                header_map = self.build_header_map(rows[0])
                active_header = self.find_active_header(header_map)
                for row in rows[1:]:
                    raw = self.extract_room_row(row, header_map, active_header)
                    if not raw:
                        continue
                    raw["source_file"] = file_name
                    raw["source_sheet"] = sheet.title
                    key = (raw["site_name"], raw["room_name"])
                    if key not in records:
                        records[key] = raw
                    else:
                        records[key] = self.merge_room_record(records[key], raw)

        if not records:
            return 0

        existing_rooms = {
            (room.site_name, room.room_name): room for room in db.query(MeetingRoom).all()
        }

        changed = 0
        for record in records.values():
            manager_name = record.get("manager_name")
            record["manager_user_id"] = self.resolve_manager_user_id(db, manager_name)

            room = existing_rooms.get((record["site_name"], record["room_name"]))
            if room:
                for field, value in record.items():
                    setattr(room, field, value)
            else:
                db.add(MeetingRoom(**record))
            changed += 1

        db.commit()
        return changed

    def build_header_map(self, header_row: Tuple[Any, ...]) -> Dict[str, int]:
        header_map: Dict[str, int] = {}
        for index, value in enumerate(header_row):
            if value is None:
                continue
            header_map[str(value).strip()] = index
        return header_map

    def find_active_header(self, header_map: Dict[str, int]) -> Optional[str]:
        for key in ["是否有效", "未来是否开放使用"]:
            if key in header_map:
                return key
        return None

    def extract_room_row(
        self,
        row: Tuple[Any, ...],
        header_map: Dict[str, int],
        active_header: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        room_name = self.normalize_text(self.get_row_value(row, header_map, "会议室名称"))
        site_name = self.normalize_text(self.get_row_value(row, header_map, "会议室分类"))
        location_text = self.normalize_text(self.get_row_value(row, header_map, "地点楼层"))
        if not room_name or not site_name or not location_text:
            return None

        building_name, floor_label = self.extract_location_parts(location_text)
        active_value = self.get_row_value(row, header_map, active_header) if active_header else None
        return {
            "room_name": room_name,
            "site_name": site_name,
            "location_text": location_text,
            "building_name": building_name,
            "floor_label": floor_label,
            "capacity": self.to_int(self.get_row_value(row, header_map, "容纳人数"), default=0),
            "seat_layout": self.normalize_text(self.get_row_value(row, header_map, "座席设置")),
            "manager_name": self.normalize_text(self.get_row_value(row, header_map, "保管者")),
            "manager_user_id": None,
            "sort_order": self.to_optional_int(self.get_row_value(row, header_map, "排序号")),
            "is_enabled": self.parse_enabled_value(active_value),
            "source_file": None,
            "source_sheet": None,
        }

    def merge_room_record(self, high_priority: Dict[str, Any], low_priority: Dict[str, Any]) -> Dict[str, Any]:
        merged = dict(high_priority)
        for key, value in low_priority.items():
            if key == "is_enabled":
                if merged.get(key) is None and value is not None:
                    merged[key] = value
                continue
            if self.is_blank(merged.get(key)) and not self.is_blank(value):
                merged[key] = value
        if merged.get("is_enabled") is None:
            merged["is_enabled"] = True
        return merged

    def get_row_value(self, row: Tuple[Any, ...], header_map: Dict[str, int], header: Optional[str]) -> Any:
        if not header or header not in header_map:
            return None
        index = header_map[header]
        return row[index] if index < len(row) else None

    def find_meeting_for_intent(
        self,
        db: Session,
        params: Dict[str, Any],
        current_user_id: Optional[str] = None,
    ) -> Meeting:
        meeting_id = self.normalize_text(params.get("meeting_id"))
        if meeting_id:
            meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
            if meeting:
                return meeting

        title = self.normalize_text(params.get("title"))
        if not title:
            raise MeetingServiceError("缺少目标会议标识，请提供会议标题或会议ID")

        query = db.query(Meeting).filter(Meeting.title.like(f"%{title}%"))
        organizer_user_id = self.parse_user_id(current_user_id)
        if organizer_user_id:
            query = query.filter(
                or_(Meeting.organizer_user_id == organizer_user_id, Meeting.organizer_user_id.is_(None))
            )

        start_time = self.parse_datetime_value(params.get("start_time"))
        if start_time:
            day_start = datetime.combine(start_time.date(), time.min)
            day_end = day_start + timedelta(days=1)
            query = query.filter(Meeting.start_time >= day_start, Meeting.start_time < day_end)

        candidates = query.order_by(Meeting.start_time.desc()).all()
        if not candidates:
            raise MeetingServiceError("未找到匹配的会议", status_code=404)
        if len(candidates) > 1:
            raise MeetingServiceError(
                "找到多个同名会议，请补充更精确的时间或会议ID",
                code="meeting_ambiguous",
                extra={
                    "candidates": [
                        {"id": item.id, "title": item.title, "start_time": item.start_time.isoformat()}
                        for item in candidates[:5]
                    ]
                },
            )
        return candidates[0]

    def resolve_manager_user_id(self, db: Session, manager_name: Optional[str]) -> Optional[int]:
        manager_name = self.normalize_text(manager_name)
        if not manager_name:
            return None

        user = (
            db.query(SysUser)
            .filter(
                or_(
                    SysUser.nick_name == manager_name,
                    SysUser.user_name == manager_name,
                    SysUser.alias == manager_name,
                )
            )
            .first()
        )
        return user.user_id if user else None

    def normalize_people(self, value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [self.normalize_text(item) for item in value if self.normalize_text(item)]
        if isinstance(value, tuple):
            return [self.normalize_text(item) for item in value if self.normalize_text(item)]

        text = self.normalize_text(value)
        if not text:
            return []
        parts = re.split(r"[、,，/；;\s]+", text)
        return [part for part in (self.normalize_text(item) for item in parts) if part]

    def build_summary(self, text: str) -> Optional[str]:
        normalized = self.normalize_text(text)
        if not normalized:
            return None

        sentences = [
            item.strip(" -")
            for item in re.split(r"[。！？；\n]+", normalized)
            if item and len(item.strip()) >= 4
        ]
        if not sentences:
            return normalized[:120]

        lines = ["会议要点："]
        for sentence in sentences[:3]:
            lines.append(f"- {sentence[:80]}")
        if len(sentences) > 3:
            lines.append(f"- 其余内容已保存在完整转写中，共 {len(sentences)} 段。")
        return "\n".join(lines)

    def parse_datetime_value(self, value: Any) -> Optional[datetime]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, time(hour=9))

        text = self.normalize_text(value)
        if not text:
            return None

        for fmt in [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d",
        ]:
            try:
                parsed = datetime.strptime(text, fmt)
                if fmt in {"%Y-%m-%d", "%Y/%m/%d"}:
                    return datetime.combine(parsed.date(), time(hour=9))
                return parsed
            except ValueError:
                continue

        chinese_date_match = re.search(
            r"(?:(?P<year>\d{4})年)?(?P<month>\d{1,2})月(?P<day>\d{1,2})日?",
            text,
        )
        today = datetime.now().date()
        if chinese_date_match:
            year = int(chinese_date_match.group("year") or today.year)
            target_date = date(year, int(chinese_date_match.group("month")), int(chinese_date_match.group("day")))
        elif "后天" in text:
            target_date = today + timedelta(days=2)
        elif "明天" in text:
            target_date = today + timedelta(days=1)
        elif "今天" in text:
            target_date = today
        else:
            target_date = today

        time_search_text = text
        if chinese_date_match:
            time_search_text = text[chinese_date_match.end():] or text

        time_match = re.search(
            r"(?P<period>上午|下午|晚上|中午|凌晨)?\s*(?P<hour>\d{1,2})(?:[:点时](?P<minute>\d{1,2}))?(?P<half>半)?",
            time_search_text,
        )
        hour = 9
        minute = 0
        if time_match:
            hour = int(time_match.group("hour"))
            minute = int(time_match.group("minute") or 0)
            period_text = time_match.group("period") or text
            if time_match.group("half") and time_match.group("minute") is None:
                minute = 30

        if "下午" in period_text or "晚上" in period_text:
            if hour < 12:
                hour += 12
        if "中午" in period_text and hour < 11:
            hour += 12
        if "凌晨" in period_text and hour == 12:
            hour = 0
        return datetime.combine(target_date, time(hour=hour, minute=minute))

    def parse_duration_minutes(self, value: Any) -> Optional[int]:
        if value is None or value == "":
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)

        text = self.normalize_text(value)
        if not text:
            return None

        minute_match = re.search(r"(\d+)\s*(?:min|分钟|分)", text, re.IGNORECASE)
        if minute_match:
            return int(minute_match.group(1))
        hour_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:hour|小时)", text, re.IGNORECASE)
        if hour_match:
            return int(float(hour_match.group(1)) * 60)
        if text.isdigit():
            return int(text)
        return None

    def parse_date_value(self, value: Optional[str]) -> Optional[date]:
        if not value:
            return None
        if isinstance(value, date):
            return value
        text = self.normalize_text(value)
        if not text:
            return None
        for fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def derive_title_from_text(self, text: str, default: str = "未命名会议") -> str:
        normalized = self.normalize_text(text)
        if not normalized:
            return default
        first_part = re.split(r"[。！？\n]+", normalized)[0].strip()
        return first_part[:40] if first_part else default

    def extract_location_parts(self, location_text: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
        normalized = self.normalize_text(location_text)
        if not normalized:
            return None, None

        patterns = [
            r"^(?P<building>.*?)(?P<floor>\d+F.*)$",
            r"^(?P<building>.*?)(?P<floor>负?\d+楼.*)$",
            r"^(?P<building>.*?)(?P<floor>[A-Za-z]?\d+层.*)$",
        ]
        for pattern in patterns:
            match = re.match(pattern, normalized)
            if match:
                building = self.normalize_text(match.group("building"))
                floor = self.normalize_text(match.group("floor"))
                return building or normalized, floor
        return normalized, None

    def parse_enabled_value(self, value: Any) -> bool:
        parsed = self.parse_bool(value, default=None)
        if parsed is None:
            return True
        return parsed

    def parse_bool(self, value: Any, default: Optional[bool] = None) -> Optional[bool]:
        if isinstance(value, bool):
            return value
        if value is None or value == "":
            return default
        text = self.normalize_text(value)
        if text in {"是", "启用", "开放", "1", "true", "yes"}:
            return True
        if text in {"否", "停用", "关闭", "0", "false", "no"}:
            return False
        return default

    def to_int(self, value: Any, default: int = 0) -> int:
        parsed = self.to_optional_int(value)
        return parsed if parsed is not None else default

    def to_optional_int(self, value: Any) -> Optional[int]:
        if value is None or value == "":
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        text = self.normalize_text(value)
        if not text:
            return None
        match = re.search(r"-?\d+", text)
        return int(match.group()) if match else None

    def parse_user_id(self, value: Optional[str]) -> Optional[int]:
        if value is None:
            return None
        text = str(value).strip()
        return int(text) if text.isdigit() else None

    def normalize_text(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def is_blank(self, value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip() == ""
        return False


meeting_service = MeetingService()
